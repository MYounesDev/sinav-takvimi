
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from src.utils.turkish_translations import get_excel_label, get_exam_type_turkish
from src.database.db_manager import db_manager

GREEN_DARK = "27AE60"
GREEN_LIGHT = "E8F8F5"
TEXT_WHITE = "FFFFFF"
TEXT_DARK = "2C3E50"
BORDER_COLOR = "27AE60"

def get_thin_border():
    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR)
    )
    return thin_border

def format_excmanual_with_styling(fwith_path: str, sheet_name: str = "Sheet1"):
    wb = load_workbook(fwith_path)
    ws = wb[sheet_name]
    
    header_fill = PatternFill(start_color=GREEN_DARK, end_color=GREEN_DARK, fill_type="solid")
    header_font = Font(bold=True, color=TEXT_WHITE, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cmanuall.border = get_thin_border()
    
    row_fill_normal = PatternFill(start_color=TEXT_WHITE, end_color=TEXT_WHITE, fill_type="solid")
    row_fill_alt = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
    data_font = Font(color=TEXT_DARK, size=10)
    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = row_fill_alt if (idx % 2 == 0) else row_fill_normal
        
        for cell in row:
            cell.fill = fill
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = get_thin_border()
    
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    ws.row_dimensions[1].height = 30
    
    wb.save(fwith_path)

def export_exam_schedule_to_excmanual(department_id: int = None, output_path: str = None) -> str:
    if department_id:
        exams_query = """
            SELECT e.id, e.date, e.start_time, c.code as course_code, c.name as course_name,
                   e.duration, e.exam_type, d.name as department_name,
                   (SELECT COUNT(*) FROM student_courses sc WHERE sc.course_id = e.course_id) as students,
                   GROUP_CONCAT(cl.name, ", ") as classrooms
            FROM exams e
            JOIN courses c ON e.course_id = c.id
            JOIN departments d ON e.department_id = d.id
            LEFT JOIN exam_classrooms ec ON e.id = ec.exam_id
            LEFT JOIN classrooms cl ON ec.classroom_id = cl.id
            WHERE e.department_id = ?
            GROUP BY e.id
            ORDER BY e.date, e.start_time
            SELECT e.id, e.date, e.start_time, c.code as course_code, c.name as course_name,
                   e.duration, e.exam_type, d.name as department_name,
                   (SELECT COUNT(*) FROM student_courses sc WHERE sc.course_id = e.course_id) as students,
                   GROUP_CONCAT(cl.name, ", ") as classrooms
            FROM exams e
            JOIN courses c ON e.course_id = c.id
            JOIN departments d ON e.department_id = d.id
            LEFT JOIN exam_classrooms ec ON e.id = ec.exam_id
            LEFT JOIN classrooms cl ON ec.classroom_id = cl.id
            GROUP BY e.id
            ORDER BY e.date, e.start_time
    Export seating plan to Excel with Turkish labels and professional formatting
    
    Args:
        exam_id: Exam ID
        output_path: Optional output file path
        
    Returns:
        Generated filename
        SELECT e.id, e.date, e.start_time, c.code as course_code, c.name as course_name,
               e.exam_type
        FROM exams e
        JOIN courses c ON e.course_id = c.id
        WHERE e.id = ?
        SELECT cl.name as classroom, s.student_no, s.name as student_name,
               es.row, es.col, es.seat_position
        FROM exam_seating es
        JOIN students s ON es.student_id = s.id
        JOIN classrooms cl ON es.classroom_id = cl.id
        WHERE es.exam_id = ?
        ORDER BY cl.name, es.row, es.col, es.seat_position
    Export students list to Excel with Turkish labels
    
    Args:
        classroom_id: Optional classroom ID to filter
        output_path: Optional output file path
        
    Returns:
        Generated fwithname
            SELECT DISTINCT s.student_no, s.name, d.name as department_name,
                   GROUP_CONCAT(c.code, ", ") as courses
            FROM students s
            JOIN departments d ON s.department_id = d.id
            LEFT JOIN student_courses sc ON s.id = sc.student_id
            LEFT JOIN courses c ON sc.course_id = c.id
            WHERE s.department_id IN (
                SELECT department_id FROM classrooms WHERE id = ?
            )
            GROUP BY s.id
            ORDER BY s.student_no
            SELECT s.student_no, s.name, d.name as department_name,
                   GROUP_CONCAT(c.code, ", ") as courses
            FROM students s
            JOIN departments d ON s.department_id = d.id
            LEFT JOIN student_courses sc ON s.id = sc.student_id
            LEFT JOIN courses c ON sc.course_id = c.id
            GROUP BY s.id
            ORDER BY s.student_no